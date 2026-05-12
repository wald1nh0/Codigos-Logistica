import pandas as pd
import requests
import time
import sys
import base64
import os
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font # <-- Nuevo para colores

# --- 1. TUS CREDENCIALES ---
BLUE_CLIENT_ID = 'Pon aqui tu Client ID de Blue'    
BLUE_CLIENT_SECRET = 'Pon aqui tu Client Secret de Blue'
BLUE_API_KEY = 'Pon aqui tu API Key de Blue'

ZIPNOVA_USER = 'Pon aqui tu usuario de ZipNova'
ZIPNOVA_PASS = 'Pon aqui tu contraseña de ZipNova'

rutStarken = 'Aqui va tu RUT de Starken sin puntos ni guion, solo números y letra final en mayúscula'
keyStarken = 'Aqui va tu clave de Starken'

# Nombres de Archivos
ARCHIVO_CONSOLIDADO = "Pon aquí el nombre de tu archivo consolidado.xlsx"
ARCHIVO_MATRIZ = "Matriz de Transito.xlsx"
HOJA_DIARIA = "EDITABLE"

URL_BLUE_TOKEN = 'https://sso.blue.cl/oauth2/token'
URL_BLUE_TRACKING = "https://cmkin.api.blue.cl/cmkin/bff/tracking-pull-corp/v1/"
URL_ZIPNOVA_BASE = "https://api.zipnova.cl/v2/shipments"
URL_STARKEN_TRACKING = "https://restservices-qa.starken.cl/apiqa/starkenservices/rest/getDetalleSeguimientoNuevo"

# VÍA 1: Tienen el ID largo (ej: 0999-12345-0001). Se limpia y va a Zipnova directo.
MARKET_ZIPNOVA = ['Dimarsa','Travel','Meli - ZipNova','Shopi - ZipNova','SAC - Bluex']

# VÍA 2: Tienen seguimiento Starken. Intentará consultar la web de Starken.
MARKET_STARKEN = ['Meli - Starken','Shopi - Starken','SAC - Starken']

# VÍA 3: Tienen seguimiento Blue. Van directo a la API de Blue Express.
MARKET_BLUE = ['Hites','Meli - Blue','Shopi - Blue']

# --- 3. FUNCIONES AUXILIARES ---
def normalizar(texto):
    if not isinstance(texto, str): return ""
    texto = texto.upper().strip()
    replacements = (("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U"))
    for a, b in replacements: texto = texto.replace(a, b)
    return texto

def normalizarIdZipNova(id_sucio):
    if not id_sucio or str(id_sucio) == "nan": return ""
    return str(id_sucio).replace('0999-', '').replace('-0001', '').strip()

def obtener_token_blue():
    print(f"-> Obteniendo Token Blue... \nextraccion actual:{path}")
    payload = {'client_id': BLUE_CLIENT_ID, 'client_secret': BLUE_CLIENT_SECRET, 'grant_type': 'client_credentials'}
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    try:
        r = requests.post(URL_BLUE_TOKEN, data=payload, headers=headers)
        if r.status_code == 200: return r.json()['access_token']
    except: pass
    print("Error obteniendo Token Blue. Saliendo."); sys.exit()

# --- 4. FUNCIONES DE CONSULTA ---

def consultar_blue(tracking, token):
    if not tracking or str(tracking) == "nan": return None
    headers = {"Authorization": f"Bearer {token}", "x-api-key": BLUE_API_KEY}
    r = requests.get(f"{URL_BLUE_TRACKING}{tracking}", headers=headers)
    
    info = {"estado": "Sin Info", "fecha_entrega": None, "fecha_recepcion": None, "comuna_real": ""}
    
    if r.status_code == 200:
        data = r.json()
        if 'deliveryAddress' in data and 'communeDesc' in data['deliveryAddress']:
            info['comuna_real'] = data['deliveryAddress']['communeDesc']
            
        info['estado'] = data.get('stateDesc', 'Sin Info')
        
        if 'packages' in data and len(data['packages']) > 0:
            paquete = data['packages'][0]
            
            # 1. BÚSQUEDA DE FECHA DE RECEPCIÓN (Ignorar digitalización)
            for evento in paquete.get('trackings', []):
                codigo = evento.get('eventCode', '')
                # Si el evento no es "Guía en Digitación" ni "Guía Enviada", es el primer scan físico
                if codigo not in ['FI', 'GE']:
                    f_rec = evento.get('eventDate')
                    if f_rec:
                        try: info['fecha_recepcion'] = pd.to_datetime(f_rec).replace(tzinfo=None)
                        except: pass
                        break # Encontramos el primero, salimos del bucle
            
            # 2. BÚSQUEDA DEL ESTADO ACTUAL REAL Y ENTREGA
            eventos = paquete.get('trackings', [])
            for evento in reversed(eventos):
                desc = evento.get('eventCodeDesc')
                if desc:
                    info['estado'] = desc
                    break
                    
            if 'latestStatus' in paquete:
                if paquete['latestStatus'].get('statusCode', '') == 'DL':
                    info['estado'] = 'ENTREGADO'
                    f_str = paquete['latestStatus'].get('statusDate')
                    if f_str:
                        try: info['fecha_entrega'] = pd.to_datetime(f_str).replace(tzinfo=None)
                        except: pass
    return info

def consultar_zipnova_directo(tracking_sucio, headers):
    id_interno = normalizarIdZipNova(tracking_sucio)
    if not id_interno: return {"estado": "ID Inválido", "fecha_entrega": None, "fecha_recepcion": None, "comuna_real": ""}

    url_track = f"{URL_ZIPNOVA_BASE}/{id_interno}/tracking?sort=newest"
    r = requests.get(url_track, headers=headers)
    info = {"estado": "Sin Info", "fecha_entrega": None, "fecha_recepcion": None, "comuna_real": ""}
    
    if r.status_code == 200:
        data = r.json()
        if isinstance(data, list) and len(data) > 0:
            info['estado'] = data[0].get('status', {}).get('name', 'Sin Info')
            if "ENTREGADO" in str(info['estado']).upper():
                try: info['fecha_entrega'] = pd.to_datetime(data[0]['created_at']).replace(tzinfo=None)
                except: pass
            
            # BÚSQUEDA DE RECEPCIÓN EN ZIPNOVA (De más antiguo a más nuevo)
            for evt in reversed(data):
                est_evt = str(evt.get('status', {}).get('name', '')).upper()
                # Ignoramos la creación, buscamos el primer movimiento
                if est_evt not in ['CREADO', 'IMPRESO', 'DOCUMENTADO', '']:
                    try: info['fecha_recepcion'] = pd.to_datetime(evt['created_at']).replace(tzinfo=None)
                    except: pass
                    break

            # Consultar Comuna
            try:
                r2 = requests.get(f"{URL_ZIPNOVA_BASE}/{id_interno}", headers=headers)
                if r2.status_code == 200:
                    d2 = r2.json()
                    if 'destination' in d2 and 'city' in d2['destination']:
                        info['comuna_real'] = d2['destination']['city']
            except: pass
    return info

def consultar_starken(tracking):
    if not tracking or str(tracking) == "nan": return None
    headers = {"Rut": rutStarken, "Clave":keyStarken, "Content-Type": "application/json"}
    payload = {"ordenFlete": tracking}
    r = requests.get(f"{URL_STARKEN_TRACKING}", headers=headers, json=payload)
    # La tubería de Starken sigue lista para cuando consigas la API
    return {"estado": "Starken: Pendiente API", "fecha_entrega": None, "fecha_recepcion": None, "comuna_real": ""}

# --- 5. PROCESO PRINCIPAL ---

print("--- ACTUALIZADOR OTIF + ALERTAS DE COLORES ---")

# A. Cargar Matriz
if not os.path.exists(ARCHIVO_MATRIZ): print(f"Falta {ARCHIVO_MATRIZ}"); sys.exit()
df_matriz = pd.read_excel(ARCHIVO_MATRIZ)
sla_dict = {normalizar(str(r.get('Comuna', ''))): {'blue': r.get('Dias_Blue', 99), 'starken': r.get('Dias_Starken', 99)} for _, r in df_matriz.iterrows()}

# B. Cargar Consolidado
df_base = pd.read_excel(ARCHIVO_CONSOLIDADO) if os.path.exists(ARCHIVO_CONSOLIDADO) else pd.DataFrame()

# C. Cargar Plantilla
root = tk.Tk(); root.withdraw()
path = filedialog.askopenfilename(title="Plantilla Diaria", filetypes=[("Excel", "*.xlsm *.xlsx")])
if not path: sys.exit()

df_nuevo = pd.read_excel(path, sheet_name=HOJA_DIARIA)
df_nuevo = df_nuevo[['Market','OPL','OC', 'SEG','SKU','PRODUCTO','Unidades','Bultos', 'Fecha Compra']]

# D. Unir Datos
if not df_base.empty:
    existentes = df_base['SEG'].astype(str).tolist()
    df_final = pd.concat([df_base, df_nuevo[~df_nuevo['SEG'].astype(str).isin(existentes)]], ignore_index=True)
else: df_final = df_nuevo

# E. Limpiar y Preparar
MARKETS_VALIDOS = MARKET_ZIPNOVA + MARKET_STARKEN + MARKET_BLUE
df_final['Market'] = df_final['Market'].astype(str).str.strip()
df_final = df_final[df_final['Market'].isin(MARKETS_VALIDOS)]

for col in ['Estado_Actual', 'Fecha_Recepcion_Courier', 'Fecha_Entrega_Real', 'Dias_Transcurridos', 'OTIF_Status', 'Comuna_Courier']:
    if col not in df_final.columns: df_final[col] = ""

# F. BARRIDO DE APIS
token_blue = obtener_token_blue()
auth_zip = base64.b64encode(f"{ZIPNOVA_USER}:{ZIPNOVA_PASS}".encode()).decode()
headers_zip = {'Authorization': f'Basic {auth_zip}', 'Content-Type': 'application/json'}

total = len(df_final)
df_final.reset_index(drop=True, inplace=True)
hoy_fecha = datetime.now() # Para medir envíos en tránsito

for i, row in df_final.iterrows():
    mkt = str(row['Market']).strip()
    track = str(row['SEG']).strip()
    est = str(row['Estado_Actual']).upper()
    
    if "ENTREGADO" in est or "CANCELADO" in est or "NULO" in est: continue
    if not track or track == "nan": continue

    data = None
    es_blue = False

    if mkt in MARKET_ZIPNOVA:
        print(f"[{i+1}/{total}] ZIPNOVA: {track}", end="\r")
        data = consultar_zipnova_directo(track, headers_zip)
    elif mkt in MARKET_STARKEN:
        data = consultar_starken(track)
    elif mkt in MARKET_BLUE:
        print(f"[{i+1}/{total}] BLUE API: {track}", end="\r")
        data = consultar_blue(track, token_blue)
        es_blue = True

    if not data: continue

    # >>> LÓGICA DE TIEMPOS Y OTIF <<<
    estado_api = str(data['estado']).upper()
    df_final.at[i, 'Estado_Actual'] = data['estado']
    if data['comuna_real']: df_final.at[i, 'Comuna_Courier'] = data['comuna_real']
    
    # 1. ¿Desde cuándo medimos? (Si no hay recepción, usamos la compra como plan B)
    if data['fecha_recepcion']:
        fecha_inicio = data['fecha_recepcion']
        df_final.at[i, 'Fecha_Recepcion_Courier'] = fecha_inicio
    else:
        fecha_inicio = pd.to_datetime(row.get('Fecha Compra')).replace(tzinfo=None) if pd.notnull(row.get('Fecha Compra')) else hoy_fecha

    # 2. Buscar Meta SLA
    comuna = normalizar(data['comuna_real']) or normalizar(str(row.get('Comuna', '')))
    meta_sla = sla_dict.get(comuna, {}).get('blue' if es_blue else 'starken', 99)

    # 3. Clasificación Semántica (Cancelado, Entregado, Tránsito)
    if "CANCELADO" in estado_api or "NULO" in estado_api:
        df_final.at[i, 'OTIF_Status'] = "Cancelado"
        
    elif "ENTREGADO" in estado_api:
        if data['fecha_entrega']: df_final.at[i, 'Fecha_Entrega_Real'] = data['fecha_entrega']
        dias = max(0, ((data['fecha_entrega'] or hoy_fecha) - fecha_inicio).days)
        df_final.at[i, 'Dias_Transcurridos'] = dias
        df_final.at[i, 'OTIF_Status'] = "Entregado" # Se pintará verde
        
    else:
        # Está en Tránsito: Calculamos con la fecha de hoy
        dias = max(0, (hoy_fecha - fecha_inicio).days)
        df_final.at[i, 'Dias_Transcurridos'] = dias
        
        if dias <= meta_sla:
            df_final.at[i, 'OTIF_Status'] = "En Tránsito (A Tiempo)" # Amarillo
        else:
            df_final.at[i, 'OTIF_Status'] = f"En Tránsito (Demorado +{dias-meta_sla}d)" # Naranja


# G. GUARDADO Y COLOREADO EN EXCEL
print(f"\n\nGuardando datos y aplicando semáforo de colores...")
df_final.to_excel(ARCHIVO_CONSOLIDADO, index=False)

try:
    wb = load_workbook(ARCHIVO_CONSOLIDADO)
    ws = wb.active
    
    # Aplicar formato de tabla base
    tab = Table(displayName="TablaOTIF", ref=ws.dimensions)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)
    
    # Encontrar la columna del OTIF_Status para pintarla
    columna_otif = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'OTIF_Status':
            columna_otif = idx
            break

    # Aplicar Colores si encontramos la columna
    if columna_otif:
        # Definición de Estilos Visuales
        color_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        font_verde = Font(color="006100", bold=True)
        
        color_amarillo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        font_amarillo = Font(color="9C6500", bold=True)
        
        color_naranja = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        font_naranja = Font(color="9C0006", bold=True)
        
        color_negro = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        font_blanca = Font(color="FFFFFF", bold=True)

        # Recorrer filas y pintar según la palabra clave
        for fila in range(2, ws.max_row + 1):
            celda = ws.cell(row=fila, column=columna_otif)
            texto_estado = str(celda.value).upper()
            
            if "ENTREGADO" in texto_estado:
                celda.fill = color_verde
                celda.font = font_verde
            elif "A TIEMPO" in texto_estado:
                celda.fill = color_amarillo
                celda.font = font_amarillo
            elif "DEMORADO" in texto_estado:
                celda.fill = color_naranja
                celda.font = font_naranja
            elif "CANCELADO" in texto_estado or "NULO" in texto_estado:
                celda.fill = color_negro
                celda.font = font_blanca

    # Auto-ajustar ancho de columnas para que se lea bien
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value)) for c in col) + 2

    wb.save(ARCHIVO_CONSOLIDADO)
    print("¡Semáforo aplicado con éxito!")
except Exception as e:
    print(f"Error aplicando colores: {e}")

print("\nLISTO. Proceso terminado.")